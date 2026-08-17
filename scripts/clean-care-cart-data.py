"""
Cleans the real Care Cart / meals log into normalized JSON shaped to match
lib/types/house-ops.ts's MealService and CareCartLog.

Reads from  ../DATA/LAF Programs 2026 - Lastest - Aug 16.xlsx   (sibling of this repo)
Writes to   ../DATA/clean/meal-services.json + care-cart-logs.json + care-cart-report.md

Both input and output live entirely outside the git repository. This script
contains only transformation logic -- no operational data -- so it is safe
to commit.

Scope: the "Care Cart" sheet only. It is NOT one table -- it's three
independent, NOT row-aligned logical tables sharing one sheet, confirmed by
direct inspection:
  1. A main day-log (cols 0-30, header row 1, data from row 3): a Google-Form-
     style daily submission with Timestamp + DATE, meal headcounts, per-time-
     slot Care Cart headcounts, activity-center counts, etc. 83 real rows,
     Mar 30 - Aug 14 2026.
  2. A "CARE CART (10:00 AM, 2:00 PM)" food-distribution ledger (cols 32-37,
     its own row-2 sub-header): DATE, FOOD DISTRIBUTED, pack size, costing,
     headcount. 32 real rows, Jan 30 - Aug 14 2026 -- a completely different
     date range than table 1, confirming these were never meant to be
     joined row-by-row.
  3. An "OPD LUNCH (12:00 NN)" ledger (cols 40-45, same shape as #2). 14 real
     rows, Mar 26 - Jul 31 2026.

MealService comes from table 1 (cols 1/3/4/5: DATE/BREAKFAST/LUNCH/DINNER) --
the only source with real per-day, per-meal-type headcounts. `costPerHead`
has no source column; MealService.costPerHead was changed from required to
optional (lib/types/house-ops.ts) rather than fabricating a peso figure.

CareCartLog comes from tables 2 and 3 -- NOT table 1's per-slot headcount
columns, which lack any item-description text (`itemsServed` is required and
not fabricatable). Table 2's own header names a combined "10:00 AM, 2:00 PM"
window rather than splitting per row, so CareCartLog.timeSlot gained a new
literal `"10:00 & 14:00"` (lib/types/house-ops.ts) rather than guessing which
of the two slots a given row belongs to. `source` ("LAF Pantry" | "Donation")
has no source column either -- also changed from required to optional.

ActivitySession is explicitly OUT OF SCOPE this pass: the sheet has
participant/volunteer counts and a free-text name list, but no session
title and no single "facilitator" field -- importing it would mean either
fabricating a title or overloading `facilitator` with a multi-name list, both
worse than leaving it as seeded mock data for now.

Usage: python scripts/clean-care-cart-data.py   (run from the repo root)
"""

import json
from datetime import datetime
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = REPO_ROOT.parent / "DATA" / "LAF Programs 2026 - Lastest - Aug 16.xlsx"
OUT_DIR = REPO_ROOT.parent / "DATA" / "clean"

# 0-based column indices, confirmed by direct inspection of the sheet's row 1/2 headers.
COL_TIMESTAMP = 0
COL_DATE = 1
COL_BREAKFAST = 3
COL_LUNCH = 4
COL_DINNER = 5

COL_SIDE1_DATE = 32
COL_SIDE1_FOOD = 33
COL_SIDE1_PACK = 34
COL_SIDE1_COUNT = 37

COL_SIDE2_DATE = 40
COL_SIDE2_FOOD = 41
COL_SIDE2_PACK = 42
COL_SIDE2_COUNT = 45

ROW_WIDTH = 46


def is_excel_error(value) -> bool:
    return isinstance(value, str) and value.startswith("#")


def num_or_none(value) -> float | None:
    if value is None or is_excel_error(value):
        return None
    if isinstance(value, (int, float)):
        return value
    return None


def iso_date(value) -> str | None:
    if value is None or is_excel_error(value):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    return text or None


def clean_str(value) -> str | None:
    if value is None or is_excel_error(value):
        return None
    text = str(value).strip()
    return text or None


def clean_meal_services(ws) -> tuple[list[dict], int, int]:
    padded = [(row + (None,) * ROW_WIDTH)[:ROW_WIDTH] for row in ws.iter_rows(min_row=3, values_only=True)]

    # Some dates have 2 submissions (resubmission/correction) -- keep only the
    # latest by Timestamp rather than emitting both or arbitrarily picking one.
    latest_by_date: dict[str, tuple] = {}
    for row in padded:
        date = iso_date(row[COL_DATE])
        if date is None:
            continue
        ts = row[COL_TIMESTAMP]
        existing = latest_by_date.get(date)
        if existing is None or (isinstance(ts, datetime) and isinstance(existing[COL_TIMESTAMP], datetime) and ts > existing[COL_TIMESTAMP]):
            latest_by_date[date] = row

    duplicate_dates_resolved = sum(
        1 for row in padded if iso_date(row[COL_DATE]) and latest_by_date.get(iso_date(row[COL_DATE])) is not row
    )

    services = []
    error_cells_skipped = 0
    for date, row in sorted(latest_by_date.items()):
        for meal_type, col in (("breakfast", COL_BREAKFAST), ("lunch", COL_LUNCH), ("dinner", COL_DINNER)):
            headcount = num_or_none(row[col])
            if headcount is None:
                if is_excel_error(row[col]):
                    error_cells_skipped += 1
                continue
            services.append({
                "date": date,
                "mealType": meal_type,
                "headcount": int(headcount),
                "exceptions": [],
            })

    return services, duplicate_dates_resolved, error_cells_skipped


def clean_care_cart_ledger(ws, date_col: int, food_col: int, pack_col: int, count_col: int, time_slot: str) -> list[dict]:
    entries = []
    for raw in ws.iter_rows(min_row=3, values_only=True):
        row = (raw + (None,) * ROW_WIDTH)[:ROW_WIDTH]
        date = iso_date(row[date_col])
        food = clean_str(row[food_col])
        headcount = num_or_none(row[count_col])
        if date is None or not food or headcount is None:
            continue
        pack = clean_str(row[pack_col])
        entries.append({
            "date": date,
            "timeSlot": time_slot,
            "itemsServed": f"{food} ({pack})" if pack else food,
            "headcount": int(headcount),
        })
    return entries


def main():
    if not SOURCE_XLSX.exists():
        raise SystemExit(f"Source workbook not found: {SOURCE_XLSX}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    ws = wb["Care Cart"]

    meal_services, duplicate_dates_resolved, error_cells_skipped = clean_meal_services(ws)

    ledger1 = clean_care_cart_ledger(ws, COL_SIDE1_DATE, COL_SIDE1_FOOD, COL_SIDE1_PACK, COL_SIDE1_COUNT, "10:00 & 14:00")
    ledger2 = clean_care_cart_ledger(ws, COL_SIDE2_DATE, COL_SIDE2_FOOD, COL_SIDE2_PACK, COL_SIDE2_COUNT, "12:00")
    care_cart_logs = sorted(ledger1 + ledger2, key=lambda e: e["date"])

    for i, m in enumerate(meal_services, start=1):
        m["id"] = f"meal-real-{i}"
    for i, c in enumerate(care_cart_logs, start=1):
        c["id"] = f"carecart-real-{i}"

    def write_json(filename, data):
        with open(OUT_DIR / filename, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    write_json("meal-services.json", meal_services)
    write_json("care-cart-logs.json", care_cart_logs)

    report_lines = [
        "# Care Cart / Meals Cleaning Report",
        "",
        f"Source: `{SOURCE_XLSX.name}` > `Care Cart` (three independent, non-row-aligned tables sharing one sheet)",
        "",
        "## MealService (from the main day-log table)",
        f"- Distinct dates with a submission: {len(set(m['date'] for m in meal_services))}",
        f"- Duplicate-date rows resolved by keeping the latest Timestamp: {duplicate_dates_resolved}",
        f"- Meal cells skipped for an Excel formula error (#REF!/#N/A), not zero-filled: {error_cells_skipped}",
        f"- Total MealService records written: {len(meal_services)}",
        "",
        "## CareCartLog (from the two food-distribution ledgers)",
        f"- 'CARE CART (10:00 AM, 2:00 PM)' ledger rows: {len(ledger1)}",
        f"- 'OPD LUNCH (12:00 NN)' ledger rows: {len(ledger2)}",
        f"- Total CareCartLog records written: {len(care_cart_logs)}",
        "",
        "## Type changes required for this data (all additive/loosening, non-breaking)",
        "- `MealService.costPerHead`: required -> optional. No source column gives a per-meal cost;",
        "  real records omit it rather than fabricating a peso figure.",
        "- `CareCartLog.timeSlot`: gained a new literal `\"10:00 & 14:00\"`. The source ledger's own header",
        "  names a combined window rather than splitting rows per slot -- forcing a guess at \"10:00\" or",
        "  \"14:00\" per row would misrepresent data the source itself doesn't distinguish.",
        "- `CareCartLog.source` (\"LAF Pantry\" | \"Donation\"): required -> optional. No source column in",
        "  either ledger distinguishes pantry stock from a donation.",
        "",
        "## Not fabricated / explicitly out of scope this pass",
        "- ActivitySession: the sheet has participant/volunteer counts and a free-text name list, but no",
        "  session title and no single facilitator field. Left as seeded mock data rather than fabricating",
        "  a title or overloading `facilitator` with a multi-name list.",
        "- The main table's own per-time-slot Care Cart headcount columns (10:00 AM / 12:00 NN / 2:00 PM /",
        "  5:00 PM) were NOT imported as CareCartLog records -- they have a real headcount per slot but no",
        "  item-description text, and `itemsServed` is a required field. A future type change (making",
        "  `itemsServed` optional) could unlock these ~83 rows of genuinely per-slot-granular headcount data.",
        "- The main table's HOUSING ACCOMMODATION column and its `#REF!` cells are occupancy-adjacent data,",
        "  left for the Occupancy integration step rather than duplicating a parallel source here.",
    ]
    (OUT_DIR / "care-cart-report.md").write_text("\n".join(report_lines), encoding="utf-8")

    print(f"Wrote {len(meal_services)} meal services and {len(care_cart_logs)} care cart logs to {OUT_DIR}")
    print(f"See {OUT_DIR / 'care-cart-report.md'} for the full report.")


if __name__ == "__main__":
    main()
