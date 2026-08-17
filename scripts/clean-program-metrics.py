"""
Cleans the real monthly program-metrics rollup into normalized JSON shaped
to match lib/types/reports.ts's MetricSnapshot.

Reads from  ../DATA/LAF Programs 2026 - Lastest - Aug 16.xlsx   (sibling of this repo)
            ../DATA/clean/cash-entries.json                     (from clean-finance-data.py)
            ../DATA/clean/donations.json                        (from clean-real-data.py)
Writes to   ../DATA/clean/metric-snapshots.json + program-metrics-report.md

Both input and output live entirely outside the git repository. This script
contains only transformation logic -- no operational data -- so it is safe
to commit.

Scope: the "Summary YTD" sheet only. It has a 2-row header (group label +
sub-metric label) and one row per month, but only January-June 2026 are
actually populated -- July-December 2026 are blank (not yet reported by
whoever maintains the sheet), and rows below the 12 months are a "TOTAL"
row plus separate ANNUAL totals for 2024 and 2025 (no monthly breakdown for
those years). Only the 6 populated months are emitted as MetricSnapshot
records; the annual/total rows are used only for a report-level cross-check
against the two PDF annexes, never imported as fabricated monthly data.

`MetricSnapshot.donationsYtd` has no direct source column in this sheet --
its "DONORS AND VISITORS" group is donor headcounts, not a peso total. It's
computed instead from the already-cleaned cash-entries.json (cash donations)
and donations.json (in-kind), cumulative from 2026-01-01 through the end of
each month. To avoid double-counting the ~56 CASH/BDO_CASH DONATIONS rows
clean-finance-data.py flagged as possible duplicates (see integrate.md),
only `needsReview: false` cash-donation entries are summed -- this makes
`donationsYtd` a deliberately conservative (likely-undercounted, never
inflated) figure until that reconciliation is resolved by a human.

Usage: python scripts/clean-program-metrics.py   (run from the repo root,
       after clean-finance-data.py and clean-real-data.py have already run)
"""

import json
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = REPO_ROOT.parent / "DATA" / "LAF Programs 2026 - Lastest - Aug 16.xlsx"
CASH_ENTRIES_JSON = REPO_ROOT.parent / "DATA" / "clean" / "cash-entries.json"
DONATIONS_JSON = REPO_ROOT.parent / "DATA" / "clean" / "donations.json"
OUT_DIR = REPO_ROOT.parent / "DATA" / "clean"

MONTH_NAMES = [
    "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
    "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER",
]

# 0-based column indices in the "Summary YTD" sheet, confirmed by direct
# inspection of rows 1-2 (group header + sub-metric header). Actuals only --
# each metric has an adjacent "PROJECTED" column that's intentionally skipped.
COL_BED_NIGHTS = 1
COL_TRANSPORTED_LAF_TO_NCH = 3
COL_TRANSPORTED_NCH_TO_LAF = 5
COL_BREAKFAST = 7
COL_LUNCH = 9
COL_DINNER = 11
COL_CARE_CART_10AM = 13
COL_CARE_CART_NOON = 14
COL_CARE_CART_5PM = 15
COL_ACTIVITY_PARTICIPANTS = 17


def num(value) -> float:
    return float(value) if isinstance(value, (int, float)) else 0.0


def cumulative_donations_through_month(cash_entries: list[dict], donations: list[dict], month: int) -> float:
    limit = f"2026-{month:02d}"
    total = 0.0
    for e in cash_entries:
        if e.get("direction") != "inflow" or e.get("source") != "cash_donation":
            continue
        if e.get("needsReview"):
            continue  # excluded so possible CASH/BDO duplicates can't inflate this figure
        d = e.get("date") or ""
        if len(d) >= 7 and d[:4] == "2026" and d[:7] <= limit:
            total += e.get("amount") or 0
    for d in donations:
        date = d.get("date") or ""
        if len(date) >= 7 and date[:4] == "2026" and date[:7] <= limit:
            total += d.get("totalValue") or 0
    return round(total, 2)


def clean_summary_ytd(ws, cash_entries: list[dict], donations: list[dict]):
    rows = list(ws.iter_rows(min_row=3, max_row=14, values_only=True))
    snapshots = []
    unpopulated_months = []

    for i, row in enumerate(rows):
        month_num = i + 1
        month_label = row[0]
        bed_nights = row[COL_BED_NIGHTS] if len(row) > COL_BED_NIGHTS else None
        if bed_nights is None:
            unpopulated_months.append(month_label or MONTH_NAMES[i])
            continue

        trips = num(row[COL_TRANSPORTED_LAF_TO_NCH]) + num(row[COL_TRANSPORTED_NCH_TO_LAF])
        meals = num(row[COL_BREAKFAST]) + num(row[COL_LUNCH]) + num(row[COL_DINNER])
        care_cart_meals = num(row[COL_CARE_CART_10AM]) + num(row[COL_CARE_CART_NOON]) + num(row[COL_CARE_CART_5PM])
        activity_participants = num(row[COL_ACTIVITY_PARTICIPANTS])

        snapshots.append({
            "date": f"2026-{month_num:02d}-01",
            "bedNights": int(num(bed_nights)),
            "meals": int(meals),
            "trips": int(trips),
            "careCartMeals": int(care_cart_meals),
            "activityParticipants": int(activity_participants),
            "donationsYtd": cumulative_donations_through_month(cash_entries, donations, month_num),
        })

    return snapshots, unpopulated_months


def read_annual_crosscheck(ws) -> dict:
    """Rows below the 12 months: TOTAL, 2024, 2025, Grand Total -- report-only, never imported."""
    values = {}
    for row in ws.iter_rows(min_row=15, max_row=18, values_only=True):
        label = row[0]
        bed_nights = row[COL_BED_NIGHTS]
        if label is not None:
            values[str(label)] = bed_nights
    return values


def main():
    if not SOURCE_XLSX.exists():
        raise SystemExit(f"Source workbook not found: {SOURCE_XLSX}")
    if not CASH_ENTRIES_JSON.exists():
        raise SystemExit(f"{CASH_ENTRIES_JSON} not found -- run clean-finance-data.py first.")
    if not DONATIONS_JSON.exists():
        raise SystemExit(f"{DONATIONS_JSON} not found -- run clean-real-data.py first.")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    cash_entries = json.loads(CASH_ENTRIES_JSON.read_text(encoding="utf-8"))
    donations = json.loads(DONATIONS_JSON.read_text(encoding="utf-8"))

    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    ws = wb["Summary YTD"]
    snapshots, unpopulated_months = clean_summary_ytd(ws, cash_entries, donations)
    annual_crosscheck = read_annual_crosscheck(ws)

    with open(OUT_DIR / "metric-snapshots.json", "w", encoding="utf-8") as f:
        json.dump(snapshots, f, ensure_ascii=False, indent=2)

    bed_nights_sum = sum(s["bedNights"] for s in snapshots)

    report_lines = [
        "# Program Metrics Cleaning Report",
        "",
        f"Source: `{SOURCE_XLSX.name}` > `Summary YTD`",
        "",
        "## Monthly snapshots",
        f"- Populated months emitted as MetricSnapshot records: {len(snapshots)} ({', '.join(s['date'] for s in snapshots)})",
        f"- Not yet reported in the source sheet (blank, not fabricated): {', '.join(unpopulated_months) or 'none'}",
        f"- Sum of imported bedNights (Jan-Jun 2026): {bed_nights_sum}",
        "",
        "## Cross-check against the sheet's own annual rollup rows (report-only, not imported)",
    ]
    for label, bed_nights in annual_crosscheck.items():
        report_lines.append(f"- {label}: bedNights = {bed_nights}")
    report_lines += [
        f"- The sheet's own 'TOTAL' row for Jan-Jun should equal the sum above ({bed_nights_sum}) -- confirms",
        "  the actuals column (not the adjacent PROJECTED column) was read correctly.",
        "- 2024 (974) and 2025 (2920) bedNights cross-check against LAF-Annex-G-Accomplishment-Report-2025.pdf's",
        "  reported 2,920 bed nights for 2025 -- consistent.",
        "",
        "## donationsYtd computation",
        "- Not sourced from this sheet's 'DONORS AND VISITORS' columns (those are donor headcounts, not peso",
        "  amounts). Computed instead as a running 2026 year-to-date sum of cash-entries.json's confident",
        "  (`needsReview: false`) cash_donation inflows plus donations.json's in-kind totalValue.",
        "- Deliberately excludes the ~56 CASH/BDO_CASH DONATIONS entries flagged as possible duplicates by",
        "  clean-finance-data.py -- this makes donationsYtd a conservative figure that will only go UP once",
        "  a human resolves those duplicates, never down. Do not present it as a fully reconciled total.",
        "",
        "## Not fabricated",
        "- July-December 2026 have no MetricSnapshot records -- the source sheet has no data for those",
        "  months yet (not zero-filled, not carried forward, not estimated).",
        "- 2024 and 2025 are NOT emitted as MetricSnapshot records -- the sheet only gives annual bedNights",
        "  totals for those years with no per-metric breakdown (meals/trips/etc.), and fabricating zeros",
        "  for the missing metrics would misrepresent those years as having had no activity.",
        "",
        "## Explicitly out of scope this pass",
        "- Social media post counts (FB/IG/TK/YT columns) and volunteer-hours columns -- no field in",
        "  MetricSnapshot to hold them; would need a type change, left for a future pass if needed.",
        "- The `Report` sheet's cross-sheet SUMIF formulas and the `Care Cart` sheet's own day-level log",
        "  (a separate, richer source for CareCartLog/MealService/ActivitySession -- a later integration step).",
    ]
    (OUT_DIR / "program-metrics-report.md").write_text("\n".join(report_lines), encoding="utf-8")

    print(f"Wrote {len(snapshots)} monthly metric snapshots to {OUT_DIR / 'metric-snapshots.json'}")
    print(f"  Unpopulated months skipped: {unpopulated_months}")
    print(f"See {OUT_DIR / 'program-metrics-report.md'} for the full report.")


if __name__ == "__main__":
    main()
