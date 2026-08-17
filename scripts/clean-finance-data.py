"""
Cleans the real LAF cash-donation and expense spreadsheets into normalized
JSON shaped to match lib/types/finance.ts's CashEntry.

Reads from  ../DATA/2026 LAF Donation Tracker.xlsx   (sibling of this repo)
Writes to   ../DATA/clean/cash-entries.json + finance-import-report.md

Both input and output live entirely outside the git repository. This script
contains only transformation logic -- no financial data -- so it is safe to
commit.

Scope: CASH, BDO_CASH DONATIONS, Bank Statement, and "Butch reimburments "
sheets only. NON-CASH (in-kind donations, already covered by
clean-real-data.py's In-kind Donations pass) and the small ad hoc stock
trackers (NON-CASH TYPE GROUP, TRACKER (EGGS, RICE), TRACKER (MEATS, FISH),
NON CASH DATABASE) are out of scope -- they're inventory data, not finance.

Two known data-quality gaps that can't be resolved by code alone, per an
explicit product decision (see integrate.md and the plan file):
  1. CASH and BDO_CASH DONATIONS overlap in date range with different
     columns. Rows are auto-merged when date+amount match AND the donor
     name matches after normalization; a date+amount collision with a
     *different* donor name is kept as two separate entries, both flagged
     `needsReview` and cross-linked via `duplicateOfId` rather than guessed.
  2. Bank Statement has no category column. Rows are classified into a
     CashEntrySource by keyword match against the free-text description;
     anything that doesn't hit a keyword is flagged `needsReview` with a
     best-guess default rather than presented as a confident classification.

Usage: python scripts/clean-finance-data.py   (run from the repo root)
"""

import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = REPO_ROOT.parent / "DATA" / "2026 LAF Donation Tracker.xlsx"
OUT_DIR = REPO_ROOT.parent / "DATA" / "clean"


def iso_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    return text or None


DATE_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")
# This workbook only covers 2025-2026 operations. A wider band (not just those
# two years) avoids flagging every row if a future year gets added later, while
# still catching the actual bugs found in this data: a plain-text non-date cell
# ("210/2026") and a real Excel date with a mistyped year (3036 instead of 2026).
PLAUSIBLE_YEAR_RANGE = (2015, 2027)


def date_is_suspect(date_str: str | None) -> str | None:
    """Returns a reason string if `date_str` looks wrong, else None. Never guess-corrects a date."""
    if date_str is None:
        return "Missing date"
    if not DATE_PATTERN.match(date_str):
        return f"Source date cell did not parse as a date (raw value: {date_str!r}) -- not corrected automatically"
    year = int(date_str[:4])
    if not (PLAUSIBLE_YEAR_RANGE[0] <= year <= PLAUSIBLE_YEAR_RANGE[1]):
        return f"Implausible year {year} in source date cell -- likely a data-entry typo, not corrected automatically"
    return None


def combine_reasons(*reasons: str | None) -> str | None:
    parts = [r for r in reasons if r]
    return "; ".join(parts) if parts else None


def clean_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def clean_amount(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def normalize_name(name: str | None) -> str:
    if not name:
        return ""
    text = re.sub(r"[^a-z0-9\s]", "", name.lower())
    return re.sub(r"\s+", " ", text).strip()


def name_similarity(a: str | None, b: str | None) -> str:
    """"exact" | "partial" | "none" -- normalized-token overlap, not a full fuzzy matcher."""
    na, nb = normalize_name(a), normalize_name(b)
    if not na or not nb:
        return "none"
    if na == nb:
        return "exact"
    ta, tb = set(na.split()), set(nb.split())
    if ta and tb and len(ta & tb) / len(ta | tb) >= 0.5:
        return "partial"
    return "none"


OUTFLOW_KEYWORDS: list[tuple[list[str], str]] = [
    (["SALARY", "PAYROLL"], "payroll"),
    (["GAS STATION", "GASOLINE", "FUEL", "PETRON", "SHELL ", "CALTEX", "UNIOIL"], "vehicle_fuel"),
    (["MERALCO", "MAYNILAD", "ELECTRIC", "WATER BILL", "PLDT", "GLOBE", "CONVERGE", " RENT"], "rent_utilities"),
    (["BANK CHARGE", "SERVICE FEE", "SVC FEE", "MAINTAINING BALANCE", "SC FEE"], "admin_ops"),
]
INFLOW_KEYWORDS: list[tuple[list[str], str]] = [
    (["DONATION", "DONOR"], "cash_donation"),
    (["GRANT"], "grant"),
    (["INTEREST"], "interest"),
]


def classify_bank_row(description: str | None, direction: str) -> tuple[str, bool]:
    """Returns (CashEntrySource, confident). `confident=False` means the value is a default guess."""
    desc_upper = (description or "").upper()
    keyword_map = OUTFLOW_KEYWORDS if direction == "outflow" else INFLOW_KEYWORDS
    default = "admin_ops" if direction == "outflow" else "cash_donation"
    for keywords, source in keyword_map:
        if any(k in desc_upper for k in keywords):
            return source, True
    return default, False


REIMBURSEMENT_CATEGORY_MAP: dict[str, str] = {
    "gasoline": "vehicle_fuel",
    "fuel": "vehicle_fuel",
    "housing supplies": "program_expense",
    "laf food supply": "program_expense",
    "food supply": "program_expense",
    "utilities": "rent_utilities",
    "rent": "rent_utilities",
}


def classify_reimbursement(category: str | None) -> tuple[str, bool]:
    key = (category or "").strip().lower()
    if key in REIMBURSEMENT_CATEGORY_MAP:
        return REIMBURSEMENT_CATEGORY_MAP[key], True
    return "admin_ops", False


def clean_cash_sheet(ws) -> list[dict]:
    """CASH: Column1(date), TIME, DONOR'S FULL NAME, CONTACT NUMBER, EMAIL, ADDRESS, TYPE, AMOUNT, INVOICE NUMBER, RECEIVED BY, NOTES"""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        (date, _time, donor_name, contact, email, address, type_, amount, invoice, received_by, notes) = (
            row + (None,) * 11
        )[:11]
        amount_val = clean_amount(amount)
        if amount_val is None or not donor_name:
            continue
        rows.append({
            "date": iso_date(date),
            "amount": amount_val,
            "donorName": clean_str(donor_name),
            "contact": clean_str(contact),
            "email": clean_str(email),
            "address": clean_str(address),
            "type": clean_str(type_),
            "invoice": clean_str(invoice),
            "receivedBy": clean_str(received_by),
            "notes": clean_str(notes),
        })
    return rows


def clean_bdo_sheet(ws) -> list[dict]:
    """BDO_CASH DONATIONS: Date, Amount, Donor's Name, Service Invoice No., link, link, NOTES"""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        (date, amount, donor_name, invoice, _link1, _link2, notes) = (row + (None,) * 7)[:7]
        amount_val = clean_amount(amount)
        if amount_val is None or not donor_name:
            continue
        rows.append({
            "date": iso_date(date),
            "amount": amount_val,
            "donorName": clean_str(donor_name),
            "invoice": clean_str(invoice),
            "notes": clean_str(notes),
        })
    return rows


def merge_cash_and_bdo(cash_rows: list[dict], bdo_rows: list[dict]) -> tuple[list[dict], int, int]:
    """Returns (merged_entries_as_partial_dicts, merged_count, flagged_pair_count).

    Matches on exact (date, amount). A single unambiguous candidate with an
    exact normalized-name match is merged into one CASH-sourced entry. A
    single candidate with a partial/no name match, or more than one
    candidate on either side, is left as separate entries and flagged
    `needsReview` rather than guessed.
    """
    from collections import defaultdict

    cash_by_key: dict[tuple, list[dict]] = defaultdict(list)
    for r in cash_rows:
        r["_matched"] = False
        cash_by_key[(r["date"], r["amount"])].append(r)

    entries = []
    merged_count = 0
    flagged_pairs = 0

    for bdo in bdo_rows:
        candidates = [c for c in cash_by_key.get((bdo["date"], bdo["amount"]), []) if not c["_matched"]]
        if len(candidates) == 1:
            cash = candidates[0]
            sim = name_similarity(cash["donorName"], bdo["donorName"])
            if sim == "exact":
                cash["_matched"] = True
                desc = f"{cash['donorName']} — {cash['type'] or 'cash'} donation"
                if cash["invoice"]:
                    desc += f" (Inv# {cash['invoice']})"
                if bdo["invoice"]:
                    desc += f" (BDO SI# {bdo['invoice']})"
                date_issue = date_is_suspect(cash["date"])
                entries.append({
                    "date": cash["date"],
                    "amount": cash["amount"],
                    "donorName": cash["donorName"],
                    "description": desc,
                    "sourceSheet": "CASH+BDO_CASH DONATIONS",
                    "needsReview": bool(date_issue),
                    "reviewReason": date_issue,
                })
                merged_count += 1
                continue
            # Same date+amount, different-looking donor name -- don't guess.
            cash["_matched"] = True
            flagged_pairs += 1
            cash_id_hint = f"cash-{cash['date']}-{cash['amount']}"
            entries.append({
                "date": cash["date"],
                "amount": cash["amount"],
                "donorName": cash["donorName"],
                "description": f"{cash['donorName']} — {cash['type'] or 'cash'} donation",
                "sourceSheet": "CASH",
                "needsReview": True,
                "reviewReason": combine_reasons(
                    f"Same date+amount as a BDO_CASH DONATIONS row for donor '{bdo['donorName']}' "
                    f"(name similarity: {sim}) -- possible duplicate, not auto-merged.",
                    date_is_suspect(cash["date"]),
                ),
                "_pairKey": cash_id_hint,
            })
            entries.append({
                "date": bdo["date"],
                "amount": bdo["amount"],
                "donorName": bdo["donorName"],
                "description": f"{bdo['donorName']} — BDO bank deposit"
                + (f" (SI# {bdo['invoice']})" if bdo["invoice"] else ""),
                "sourceSheet": "BDO_CASH DONATIONS",
                "needsReview": True,
                "reviewReason": combine_reasons(
                    f"Same date+amount as a CASH row for donor '{cash['donorName']}' "
                    f"(name similarity: {sim}) -- possible duplicate, not auto-merged.",
                    date_is_suspect(bdo["date"]),
                ),
                "_pairKey": cash_id_hint,
            })
            continue
        if len(candidates) > 1:
            # Ambiguous: multiple CASH rows share this date+amount. Don't pick one.
            flagged_pairs += 1
            entries.append({
                "date": bdo["date"],
                "amount": bdo["amount"],
                "donorName": bdo["donorName"],
                "description": f"{bdo['donorName']} — BDO bank deposit"
                + (f" (SI# {bdo['invoice']})" if bdo["invoice"] else ""),
                "sourceSheet": "BDO_CASH DONATIONS",
                "needsReview": True,
                "reviewReason": combine_reasons(
                    f"{len(candidates)} CASH rows share this date+amount -- could not disambiguate automatically.",
                    date_is_suspect(bdo["date"]),
                ),
            })
            continue
        # No candidate at all -- genuinely distinct entry.
        date_issue = date_is_suspect(bdo["date"])
        entries.append({
            "date": bdo["date"],
            "amount": bdo["amount"],
            "donorName": bdo["donorName"],
            "description": f"{bdo['donorName']} — BDO bank deposit"
            + (f" (SI# {bdo['invoice']})" if bdo["invoice"] else ""),
            "sourceSheet": "BDO_CASH DONATIONS",
            "needsReview": bool(date_issue),
            "reviewReason": date_issue,
        })

    # Remaining unmatched CASH rows.
    for r in cash_rows:
        if r["_matched"]:
            continue
        desc = f"{r['donorName']} — {r['type'] or 'cash'} donation"
        if r["invoice"]:
            desc += f" (Inv# {r['invoice']})"
        date_issue = date_is_suspect(r["date"])
        entries.append({
            "date": r["date"],
            "amount": r["amount"],
            "donorName": r["donorName"],
            "description": desc,
            "sourceSheet": "CASH",
            "needsReview": bool(date_issue),
            "reviewReason": date_issue,
        })

    return entries, merged_count, flagged_pairs


def clean_bank_statement(ws) -> list[dict]:
    """Posting Date, Branch, Description, Debit, Credit, Running Balance, Check Number"""
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        (date, _branch, description, debit, credit, _balance, _check) = (row + (None,) * 7)[:7]
        debit_val = clean_amount(debit)
        credit_val = clean_amount(credit)
        if not debit_val and not credit_val:
            continue
        direction = "outflow" if debit_val else "inflow"
        amount = debit_val if debit_val else credit_val
        source, confident = classify_bank_row(description, direction)
        date_str = iso_date(date)
        date_issue = date_is_suspect(date_str)
        entries.append({
            "date": date_str,
            "direction": direction,
            "amount": amount,
            "source": source,
            "description": clean_str(description) or "(no description)",
            "sourceSheet": "Bank Statement",
            "needsReview": not confident or bool(date_issue),
            "reviewReason": combine_reasons(
                None if confident else f"No matching keyword in bank description; defaulted to '{source}' for a {direction} row.",
                date_issue,
            ),
        })
    return entries


def clean_reimbursements(ws) -> list[dict]:
    """DATE, PAYEE, TIN#, ADDRESS, DESCRIPTION, CATEGORY, LINK TO OR, GROSS AMT, Column1"""
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        (date, payee, _tin, _address, description, category, _link, gross_amt, _col) = (row + (None,) * 9)[:9]
        amount = clean_amount(gross_amt)
        if amount is None or not payee:
            continue
        source, confident = classify_reimbursement(category)
        desc = f"{payee}: {description}" if description else str(payee)
        date_str = iso_date(date)
        date_issue = date_is_suspect(date_str)
        entries.append({
            "date": date_str,
            "direction": "outflow",
            "amount": amount,
            "source": source,
            "donorName": clean_str(payee),
            "description": clean_str(desc),
            "sourceSheet": "Butch reimburments",
            "needsReview": not confident or bool(date_issue),
            "reviewReason": combine_reasons(
                None if confident else f"Category '{category}' has no known mapping; defaulted to 'admin_ops'.",
                date_issue,
            ),
        })
    return entries


def main():
    if not SOURCE_XLSX.exists():
        raise SystemExit(f"Source workbook not found: {SOURCE_XLSX}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)

    cash_rows = clean_cash_sheet(wb["CASH"])
    bdo_rows = clean_bdo_sheet(wb["BDO_CASH DONATIONS"])
    merged_donation_entries, merged_count, flagged_pairs = merge_cash_and_bdo(cash_rows, bdo_rows)
    bank_entries = clean_bank_statement(wb["Bank Statement"])
    reimbursement_entries = clean_reimbursements(wb["Butch reimburments "])

    cash_entries = []
    next_id = 1

    # Two-pass id assignment so duplicateOfId can reference the paired entry's real id.
    pair_key_to_ids: dict[str, list[str]] = {}
    for e in merged_donation_entries:
        entry_id = f"cash-real-{next_id}"
        next_id += 1
        pair_key = e.pop("_pairKey", None)
        if pair_key:
            pair_key_to_ids.setdefault(pair_key, []).append(entry_id)
        cash_entries.append({
            "id": entry_id,
            "date": e["date"],
            "direction": "inflow",
            "source": "cash_donation",
            "entity": "PH_SEC",
            "currency": "PHP",
            "amount": e["amount"],
            "description": e["description"],
            # Real historical entries already happened -- "pending" here would mean
            # "awaiting a threshold-approval decision" (a different workflow, see
            # app/(app)/finance/approvals), not "data quality needs review". Those
            # are tracked separately via needsReview/reviewReason instead.
            "approvalStatus": "approved",
            "donorName": e["donorName"],
            "sourceSheet": e["sourceSheet"],
            "needsReview": e["needsReview"],
            "reviewReason": e.get("reviewReason"),
            "_pairKey": pair_key,
        })

    for pair_key, ids in pair_key_to_ids.items():
        if len(ids) == 2:
            for entry in cash_entries:
                if entry.get("_pairKey") == pair_key:
                    entry["duplicateOfId"] = next(i for i in ids if i != entry["id"])

    for e in cash_entries:
        e.pop("_pairKey", None)

    for e in bank_entries:
        entry_id = f"cash-real-{next_id}"
        next_id += 1
        cash_entries.append({
            "id": entry_id,
            "date": e["date"],
            "direction": e["direction"],
            "source": e["source"],
            "entity": "PH_SEC",
            "currency": "PHP",
            "amount": e["amount"],
            "description": e["description"],
            # Real historical entries already happened -- "pending" here would mean
            # "awaiting a threshold-approval decision" (a different workflow, see
            # app/(app)/finance/approvals), not "data quality needs review". Those
            # are tracked separately via needsReview/reviewReason instead.
            "approvalStatus": "approved",
            "sourceSheet": e["sourceSheet"],
            "needsReview": e["needsReview"],
            "reviewReason": e["reviewReason"],
        })

    for e in reimbursement_entries:
        entry_id = f"cash-real-{next_id}"
        next_id += 1
        cash_entries.append({
            "id": entry_id,
            "date": e["date"],
            "direction": e["direction"],
            "source": e["source"],
            "entity": "PH_SEC",
            "currency": "PHP",
            "amount": e["amount"],
            "description": e["description"],
            # Real historical entries already happened -- "pending" here would mean
            # "awaiting a threshold-approval decision" (a different workflow, see
            # app/(app)/finance/approvals), not "data quality needs review". Those
            # are tracked separately via needsReview/reviewReason instead.
            "approvalStatus": "approved",
            "donorName": e["donorName"],
            "sourceSheet": e["sourceSheet"],
            "needsReview": e["needsReview"],
            "reviewReason": e["reviewReason"],
        })

    def write_json(filename, data):
        with open(OUT_DIR / filename, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    write_json("cash-entries.json", cash_entries)

    needs_review_count = sum(1 for e in cash_entries if e.get("needsReview"))
    date_issue_count = sum(
        1 for e in cash_entries if e.get("reviewReason") and "date" in e["reviewReason"].lower()
    )
    bank_flagged = sum(1 for e in bank_entries if e["needsReview"])
    reimb_flagged = sum(1 for e in reimbursement_entries if e["needsReview"])

    report_lines = [
        "# Finance Data Cleaning Report",
        "",
        f"Source: `{SOURCE_XLSX.name}`",
        "",
        "## Cash donations (CASH + BDO_CASH DONATIONS)",
        f"- CASH sheet rows read: {len(cash_rows)}",
        f"- BDO_CASH DONATIONS sheet rows read: {len(bdo_rows)}",
        f"- Auto-merged (exact date+amount+donor-name match): {merged_count}",
        f"- Flagged as possible duplicates (date+amount match, name mismatch or ambiguous): {flagged_pairs} pair(s)",
        f"- Total cash-donation entries emitted: {len(merged_donation_entries)}",
        "",
        "## Bank Statement",
        f"- Rows read: {len(bank_entries)}",
        f"- Classified by keyword with high confidence: {len(bank_entries) - bank_flagged}",
        f"- Flagged `needsReview` (no keyword match, defaulted): {bank_flagged}",
        "",
        "## Butch reimburments (expense reimbursements)",
        f"- Rows read: {len(reimbursement_entries)}",
        f"- Classified by known CATEGORY mapping: {len(reimbursement_entries) - reimb_flagged}",
        f"- Flagged `needsReview` (unmapped category, defaulted to admin_ops): {reimb_flagged}",
        "",
        "## Totals",
        f"- Total CashEntry records written: {len(cash_entries)}",
        f"- Total flagged `needsReview` (needs a human look before use in tax receipts/board reporting): {needs_review_count}",
        f"- Of which flagged for an implausible/unparseable source date (not auto-corrected): {date_issue_count}",
        "",
        "## Fields defaulted for schema compatibility (not present in source data)",
        "- `entity`: defaulted to `\"PH_SEC\"` for every entry (source has no per-entry entity split).",
        "- `currency`: defaulted to `\"PHP\"` (source amounts are all peso figures).",
        "- `programId`: omitted entirely -- no program mapping exists in the source sheets, not fabricated.",
        "- `approvalStatus`: always `\"approved\"` -- these are historical transactions that already",
        "  happened, not entries awaiting the app's threshold-approval workflow (app/(app)/finance/approvals).",
        "  Data-quality confidence is tracked separately via `needsReview`/`reviewReason` -- a human still",
        "  needs to review flagged entries before they're treated as authoritative for tax receipts or",
        "  board reporting, but they should NOT show up in the Approvals queue, which is a different concept.",
        "",
        "## Explicitly out of scope this pass",
        "- `INCOME AND EXPENSES TRACKING` (monthly rollup formulas) -- use only to spot-check totals.",
        "- `NON-CASH TYPE GROUP`, `NON CASH DATABASE`, `TRACKER (EGGS, RICE)`, `TRACKER (MEATS, FISH)` -- inventory data, not finance.",
        "- `NON-CASH` (in-kind donations) -- already covered by clean-real-data.py.",
    ]
    (OUT_DIR / "finance-import-report.md").write_text("\n".join(report_lines), encoding="utf-8")

    print(f"Wrote {len(cash_entries)} cash entries to {OUT_DIR / 'cash-entries.json'}")
    print(f"  {merged_count} merged, {flagged_pairs} flagged duplicate pair(s), "
          f"{bank_flagged} bank rows + {reimb_flagged} reimbursement rows need review")
    print(f"See {OUT_DIR / 'finance-import-report.md'} for the full report.")


if __name__ == "__main__":
    main()
