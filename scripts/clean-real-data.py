"""
Cleans the real LAF operational spreadsheets (Patients master + Donors/Donations)
into normalized JSON, shaped to closely match the app's TypeScript types
(lib/types/patient.ts, lib/types/donor.ts, lib/types/reference.ts) so a future
import is close to a direct mapping.

Reads from  ../DATA/LAF PROGRAMS 2026.xlsx        (sibling of this repo)
Writes to   ../DATA/clean/*.json + import-report.md (sibling of this repo)

Both input and output live entirely outside the git repository. This script
contains only transformation logic -- no patient data -- so it is safe to commit.

Scope: Patients master + Donors/Donations only. The 248-sheet Occupancy Tracker
(real bed-stay history) and the messy House Ops aggregate sheets (Care Cart,
REPORT, Summary YTD, monthly grids) are explicitly out of scope for this pass.

Usage: python scripts/clean-real-data.py   (run from the repo root)
"""

import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = REPO_ROOT.parent / "DATA" / "LAF PROGRAMS 2026.xlsx"
OUT_DIR = REPO_ROOT.parent / "DATA" / "clean"

# Mirrors lib/mock-data/reference-data.ts's curated lists, so we only emit a
# delta of genuinely new reference entries rather than re-declaring the ones
# that already exist in the app.
EXISTING_PROVINCES = {
    "metro manila": "prov-ncr",
    "cavite": "prov-cavite",
    "laguna": "prov-laguna",
    "bulacan": "prov-bulacan",
    "rizal": "prov-rizal",
    "batangas": "prov-batangas",
    "pampanga": "prov-pampanga",
    "quezon": "prov-quezon",
    "ncr": "prov-ncr",
}
EXISTING_DIAGNOSES = {
    "acute lymphoblastic leukemia": "dx-all",
    "acute myeloid leukemia": "dx-aml",
    "wilms tumor": "dx-wilms",
    "osteosarcoma": "dx-osteo",
    "neuroblastoma": "dx-neuro",
    "retinoblastoma": "dx-retino",
    "hodgkin lymphoma": "dx-lymphoma",
    "thalassemia major": "dx-thal-major",
    "thalassemia intermedia": "dx-thal-inter",
    "other chronic illness": "dx-other",
}
EXISTING_PHASES = {
    "diagnostic workup": "phase-diagnosis",
    "induction chemotherapy": "phase-induction",
    "consolidation": "phase-consolidation",
    "maintenance": "phase-maintenance",
    "regular transfusion": "phase-transfusion",
    "post-treatment surveillance": "phase-surveillance",
    "palliative / comfort care": "phase-palliative",
}

STATUS_MAP = {
    "on-going treatment": "ongoing",
    "expired": "expired",
    "non-pedia": "non_pedia",
    "check up": "check_up",
}

# Real data uses abbreviations for two diagnoses already in the curated list
# (EXISTING_DIAGNOSES has the spelled-out names). Without this, "ALL"/"AML"
# would slugify to `dx-all`/`dx-aml` -- which happen to be the *exact same ids*
# the curated list already uses for the spelled-out versions -- a silent id
# collision. Route abbreviations to the same canonical name so they resolve
# to the existing id instead of minting a colliding new one.
DIAGNOSIS_SYNONYMS = {
    "all": "acute lymphoblastic leukemia",
    "aml": "acute myeloid leukemia",
    "bcell all": "acute lymphoblastic leukemia",
    "tcell all": "acute lymphoblastic leukemia",
    "b-cell all": "acute lymphoblastic leukemia",
    "t-cell all": "acute lymphoblastic leukemia",
}

# Substring keywords (checked against the full lowercased name) plus a set of
# exact-token abbreviations (checked against individual words, so "all" doesn't
# false-positive-match as a substring of unrelated words).
CANCER_KEYWORDS = ["leukemia", "lymphoma", "tumor", "tumour", "sarcoma", "carcinoma", "blastoma", "malignan"]
CANCER_ABBREVIATIONS = {"all", "aml", "cml", "cll"}


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.strip().lower()).strip("-")
    return slug or "unknown"


def diagnosis_category(name: str) -> str:
    lower = name.lower()
    if "thalassemia" in lower:
        return "thalassemia"
    if any(k in lower for k in CANCER_KEYWORDS):
        return "cancer"
    tokens = set(re.findall(r"[a-z]+", lower))
    if tokens & CANCER_ABBREVIATIONS:
        return "cancer"
    return "other"


def iso_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip() or None


def clean_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def clean_phone(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def split_name(full_name: str) -> tuple[str, str]:
    if "," in full_name:
        last, first = full_name.split(",", 1)
        return first.strip(), last.strip()
    parts = full_name.strip().split(" ", 1)
    if len(parts) == 2:
        return parts[1].strip(), parts[0].strip()
    return "", full_name.strip()


def get_or_create_ref(cache: dict, existing: dict, name: str, prefix: str, synonyms: dict | None = None) -> str:
    key = name.strip().lower()
    key = (synonyms or {}).get(key, key)
    if key in existing:
        return existing[key]
    if name in cache:
        return cache[name]["id"]
    ref_id = f"{prefix}-{slugify(name)}"
    # Guard against a newly-minted id accidentally colliding with a curated
    # one whose name didn't match above (e.g. an abbreviation not covered by
    # `synonyms`) -- disambiguate rather than silently overwrite.
    existing_ids = set(existing.values())
    if any(v["id"] == ref_id for v in cache.values()) or ref_id in existing_ids:
        n = 2
        while f"{ref_id}-{n}" in existing_ids or any(v["id"] == f"{ref_id}-{n}" for v in cache.values()):
            n += 1
        ref_id = f"{ref_id}-{n}"
    cache[name] = {"id": ref_id, "name": name}
    return ref_id


def clean_patients(wb):
    ws = wb["FINAL_PATIENTS DATABASE"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    new_provinces: dict = {}
    new_diagnoses: dict = {}
    new_phases: dict = {}

    patients = []
    carers = []
    skipped = []

    for row in rows:
        (
            patient_number, date_enrolled, name, birthday, present_age, age_bracket,
            sex, address, province_city, region, patient_status, illness_type,
            diagnosis, treatment_phase, carer_name, relationship, cellphone,
            marital_status, remarks,
        ) = (row + (None,) * 19)[:19]

        if patient_number is None or not name:
            continue

        first_name, last_name = split_name(str(name))
        if not last_name:
            skipped.append({"row": name, "reason": "could not parse name"})
            continue

        status_key = str(patient_status).strip().lower() if patient_status else ""
        status = STATUS_MAP.get(status_key)
        if status is None:
            skipped.append({"row": name, "reason": f"unrecognized status {patient_status!r}"})
            continue

        patient_id = f"pt-real-{int(patient_number)}"

        province_id = None
        if province_city:
            province_id = get_or_create_ref(new_provinces, EXISTING_PROVINCES, str(province_city).strip(), "prov")

        diagnosis_ids = []
        if diagnosis:
            diagnosis_id = get_or_create_ref(
                new_diagnoses, EXISTING_DIAGNOSES, str(diagnosis).strip(), "dx", synonyms=DIAGNOSIS_SYNONYMS
            )
            diagnosis_ids.append(diagnosis_id)

        treatment_phase_id = None
        if treatment_phase:
            treatment_phase_id = get_or_create_ref(new_phases, EXISTING_PHASES, str(treatment_phase).strip(), "phase")

        patients.append({
            "id": patient_id,
            "patientNumber": str(patient_number).strip() if not isinstance(patient_number, float) else str(int(patient_number)),
            "firstName": first_name,
            "lastName": last_name,
            "birthDate": iso_date(birthday),
            "sex": clean_str(sex),
            "rawAddress": clean_str(address),
            "provinceId": province_id,
            "region": clean_str(region),
            "diagnosisIds": diagnosis_ids,
            "treatmentPhaseId": treatment_phase_id,
            "status": status,
            "carerIds": [f"carer-{patient_id}"] if carer_name else [],
            "admittedAt": iso_date(date_enrolled),
            "maritalStatus": clean_str(marital_status),
            "remarks": clean_str(remarks),
        })

        if carer_name:
            carers.append({
                "id": f"carer-{patient_id}",
                "patientId": patient_id,
                "name": str(carer_name).strip(),
                "relationship": clean_str(relationship),
                "mobileNumber": clean_phone(cellphone),
                "effectiveFrom": iso_date(date_enrolled),
            })

    delta = {
        "provinces": [
            {"id": v["id"], "name": v["name"], "region": None} for v in new_provinces.values()
        ],
        "diagnoses": [
            {"id": v["id"], "name": v["name"], "category": diagnosis_category(v["name"])} for v in new_diagnoses.values()
        ],
        "treatmentPhases": [
            {"id": v["id"], "name": v["name"]} for v in new_phases.values()
        ],
    }

    return patients, carers, delta, skipped


def clean_donors_and_donations(wb):
    donors_by_key: dict[str, dict] = {}
    donations = []

    def get_or_create_donor(raw_name, contact=None, email=None, seen_date=None):
        name = (clean_str(raw_name) or "Anonymous").strip()
        key = name.lower()
        is_anon = key == "anonymous"
        if is_anon:
            key = "__anonymous__"
            name = "Anonymous"
        if key not in donors_by_key:
            donors_by_key[key] = {
                "id": f"donor-real-{len(donors_by_key) + 1}",
                "name": name,
                "type": "anonymous" if is_anon else "individual",
                "email": clean_str(email),
                "phone": clean_phone(contact),
                "taxJurisdiction": "PH",
                "firstGiftDate": seen_date,
                "lastGiftDate": seen_date,
                "lifetimeValue": 0,
                "giftCount": 0,
            }
        else:
            d = donors_by_key[key]
            d["email"] = d["email"] or clean_str(email)
            d["phone"] = d["phone"] or clean_phone(contact)
            if seen_date:
                if not d["firstGiftDate"] or seen_date < d["firstGiftDate"]:
                    d["firstGiftDate"] = seen_date
                if not d["lastGiftDate"] or seen_date > d["lastGiftDate"]:
                    d["lastGiftDate"] = seen_date
        return donors_by_key[key]

    # Pass 1: DonorsVisitors Information -- contact records (no gift amounts).
    ws_dv = wb["DonorsVisitors Information"]
    for row in ws_dv.iter_rows(min_row=2, values_only=True):
        date, name, contact, email = (row + (None,) * 4)[:4]
        if not name:
            continue
        get_or_create_donor(name, contact=contact, email=email, seen_date=iso_date(date))

    # Pass 2: In-kind Donations -- actual gift line items.
    ws_ik = wb["In-kind Donations"]
    skipped_donations = 0
    for row in ws_ik.iter_rows(min_row=3, values_only=True):
        (
            week, date, time, donor_name, contact, email, address,
            description, item_type, qty, cost_per_unit, total_cost,
        ) = (row + (None,) * 12)[:12]

        if not description and not donor_name:
            continue

        donor = get_or_create_donor(donor_name or "Anonymous", contact=contact, email=email, seen_date=iso_date(date))
        donor["giftCount"] += 1

        qty_val = qty if isinstance(qty, (int, float)) else None
        unit_val = cost_per_unit if isinstance(cost_per_unit, (int, float)) else None
        total_val = total_cost if isinstance(total_cost, (int, float)) else (
            qty_val * unit_val if qty_val is not None and unit_val is not None else None
        )
        if total_val is None:
            skipped_donations += 1
            total_val = 0

        donor["lifetimeValue"] += total_val

        donations.append({
            "id": f"don-real-{len(donations) + 1}",
            "donorId": donor["id"],
            "date": iso_date(date),
            "receivingEntity": "PH_SEC",
            "kind": "in_kind",
            "itemDescription": clean_str(description),
            "itemType": clean_str(item_type),
            "quantity": qty_val,
            "unitValue": unit_val,
            "totalValue": total_val,
            "currency": "PHP",
        })

    donors = list(donors_by_key.values())
    return donors, donations, skipped_donations


def main():
    if not SOURCE_XLSX.exists():
        raise SystemExit(f"Source workbook not found: {SOURCE_XLSX}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)

    patients, carers, ref_delta, skipped_patients = clean_patients(wb)
    donors, donations, skipped_donations = clean_donors_and_donations(wb)

    def write_json(filename, data):
        with open(OUT_DIR / filename, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    write_json("patients.json", patients)
    write_json("carers.json", carers)
    write_json("donors.json", donors)
    write_json("donations.json", donations)
    write_json("reference-data-delta.json", ref_delta)

    report_lines = [
        "# Real Data Cleaning Report",
        "",
        f"Source: `{SOURCE_XLSX.name}`",
        "",
        "## Patients",
        f"- Cleaned: {len(patients)}",
        f"- Skipped: {len(skipped_patients)}",
    ]
    for s in skipped_patients:
        report_lines.append(f"  - {s['row']!r}: {s['reason']}")
    report_lines += [
        "",
        "## Carers",
        f"- Cleaned: {len(carers)}",
        "",
        "## Donors",
        f"- Unique donors (deduped by normalized name): {len(donors)}",
        "",
        "## Donations",
        f"- Cleaned: {len(donations)}",
        f"- Rows with no computable total value (defaulted to 0): {skipped_donations}",
        "",
        "## Reference data delta (not yet in lib/mock-data/reference-data.ts)",
        f"- New provinces: {len(ref_delta['provinces'])} -> {[p['name'] for p in ref_delta['provinces']]}",
        f"- New diagnoses: {len(ref_delta['diagnoses'])} -> {[d['name'] for d in ref_delta['diagnoses']]}",
        f"- New treatment phases: {len(ref_delta['treatmentPhases'])} -> {[p['name'] for p in ref_delta['treatmentPhases']]}",
        "",
        "## Diagnoses flagged for manual category review",
        "`category` is a best-effort keyword guess and cannot be fully trusted. Entries below matched",
        "no cancer keyword but contain \"stage\", a strong staged-malignancy signal, or another marker",
        "worth a human look before this data is treated as authoritative:",
    ] + [
        f"- `{d['id']}` {d['name']!r} (currently `{d['category']}`)"
        for d in ref_delta["diagnoses"]
        if d["category"] != "cancer" and "stage" in d["name"].lower()
    ] + [
        "",
        "## Fields defaulted for schema compatibility (not present in source data)",
        "- `Patient.isolationRequired`, `Patient.photoConsentGranted`: omitted entirely (not fabricated).",
        "- `Patient.cityId`: omitted; `rawAddress` carries the literal address string instead. The current",
        "  `lib/mock-data/reference-data.ts` city list is too sparse to fuzzy-match real addresses against.",
        "- `Donor.taxJurisdiction`: defaulted to `\"PH\"` for every donor (source has no per-donor jurisdiction).",
        "- `Donation.receivingEntity`: defaulted to `\"PH_SEC\"` for every donation (not present in source).",
        "- `Donation.currency`: defaulted to `\"PHP\"` (source amounts are all peso figures, no currency column).",
        "- Donor `firstGiftDate`/`lastGiftDate`/`lifetimeValue`/`giftCount` for donors with no matching",
        "  `In-kind Donations` row are derived only from their `DonorsVisitors Information` visit date(s),",
        "  with `lifetimeValue`/`giftCount` left at 0 -- they were contacts/visitors, not necessarily givers.",
        "",
        "## Explicitly out of scope this pass",
        "- Real bed-stay history from `LAF Occupancy Tracker.xlsx` (248 daily sheets).",
        "- `Care Cart`, `REPORT`, `Summary YTD`, and the monthly grid sheets in `LAF PROGRAMS 2026.xlsx`.",
    ]
    (OUT_DIR / "import-report.md").write_text("\n".join(report_lines), encoding="utf-8")

    print(f"Wrote {len(patients)} patients, {len(carers)} carers, {len(donors)} donors, "
          f"{len(donations)} donations to {OUT_DIR}")
    print(f"Reference delta: {len(ref_delta['provinces'])} provinces, "
          f"{len(ref_delta['diagnoses'])} diagnoses, {len(ref_delta['treatmentPhases'])} phases")
    print(f"See {OUT_DIR / 'import-report.md'} for the full report.")


if __name__ == "__main__":
    main()
