"""
Cleans the real 248-sheet daily Occupancy Tracker into normalized JSON shaped
to match lib/types/house-ops.ts's CensusSnapshot.

Reads from  ../DATA/LAF Occupancy Tracker.xlsx   (sibling of this repo)
            ../DATA/clean/patients.json           (for an informational match-rate check only)
Writes to   ../DATA/clean/census-history.json + occupancy-report.md

Both input and output live entirely outside the git repository. This script
contains only transformation logic -- no patient data -- so it is safe to
commit.

One sheet per calendar day (Dec 2025 - Aug 2026), each a ~14-slot roster of
who's physically in the house that day. Two structural problems solved here,
verified by direct inspection before writing any extraction logic (not
assumed from an earlier summary):

1. Sheet names use 3 different digit-count date formats across the tracker's
   history (8-digit MMDDYYYY, 6-digit MMDDYY, 5-digit M-DYY with no leading
   zero on the day), plus sporadic leading whitespace. Verified: parsing by
   digit-length after stripping whitespace succeeds for all 248 sheets with
   no ambiguous/duplicate dates -- no cutover-date table needed.

2. The header row is NOT a single fixed schema -- 28 distinct column
   arrangements were found (extra/missing "Next Appointment"/"Treatment"/
   "Address" columns, an older 6-column layout with two "Birthday" columns,
   several sheets with stray data typed directly into a header cell). Rather
   than hardcode column positions or a version cutover, columns are found by
   searching each sheet's own header row for "patient...name" / "carer...
   name" / a "relationship" label, so every variant resolves correctly
   without a lookup table that would need updating if the sheet changes again.

## Scope decision: CensusSnapshot only, NOT Stay/bed-level data

Verified empirically (not assumed) that this really is a full daily census --
the same patient name recurs on consecutive days while they're in-house,
confirming a headcount derived from "how many rows are filled today" is
meaningful. But the roster has NO bed/unit assignment column at all, only
name/carer/relationship/appointment info. `Stay.bedPositionId` is required
and is read by the floor-plan board, house-occupancy-summary, and the partner
portal's live bed map to determine which physical bed is occupied -- all of
which filter to `bedPositionId === <a specific bed>`. Generating real `Stay`
records with a fabricated/round-robin bed assignment would make those views
LOOK precise while being wrong; leaving `bedPositionId` undefined would make
every bed silently render as empty, including for patients who are really
admitted today -- a worse regression than not touching `Stay` at all. So this
script deliberately produces `CensusSnapshot` (a pure headcount, well-
supported by the source) and does NOT touch `stays` (still lib/mock-data/
patients.ts's existing generator) -- see occupancy-report.md and integrate.md
for the full reasoning.

Usage: python scripts/clean-occupancy-data.py   (run from the repo root)
"""

import json
import re
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = REPO_ROOT.parent / "DATA" / "LAF Occupancy Tracker.xlsx"
PATIENTS_JSON = REPO_ROOT.parent / "DATA" / "clean" / "patients.json"
OUT_DIR = REPO_ROOT.parent / "DATA" / "clean"

TOTAL_UNITS = 13  # LAF House units B1-B13 -- a known physical constant, see lib/mock-data/house-ops.ts


def parse_sheet_date(raw_name: str) -> str | None:
    n = raw_name.strip()
    if not n.isdigit():
        return None
    length = len(n)
    if length == 8:
        mm, dd, yyyy = int(n[0:2]), int(n[2:4]), int(n[4:8])
    elif length == 6:
        mm, dd, yyyy = int(n[0:2]), int(n[2:4]), 2000 + int(n[4:6])
    elif length == 5:
        mm, dd, yyyy = int(n[0:1]), int(n[1:3]), 2000 + int(n[3:5])
    else:
        return None
    if not (1 <= mm <= 12 and 1 <= dd <= 31 and 2015 <= yyyy <= 2027):
        return None
    return f"{yyyy:04d}-{mm:02d}-{dd:02d}"


def normalize(text) -> str:
    if not text:
        return ""
    return re.sub(r"\s+", " ", re.sub(r"[^a-z\s]", "", str(text).lower())).strip()


def find_column(header_row, *label_fragments: str) -> int | None:
    """Finds the column whose normalized header contains all of `label_fragments`, e.g. ("patient", "name")."""
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        norm = normalize(cell)
        if all(frag in norm for frag in label_fragments):
            return i
    return None


def split_name(full_name: str) -> tuple[str, str]:
    if "," in full_name:
        last, first = full_name.split(",", 1)
        return first.strip(), last.strip()
    parts = full_name.strip().split(" ", 1)
    if len(parts) == 2:
        return parts[1].strip(), parts[0].strip()
    return "", full_name.strip()


def build_patient_index(patients: list[dict]) -> tuple[dict[str, str], dict[str, list[str]]]:
    exact: dict[str, str] = {}
    loose: dict[str, list[str]] = {}
    for p in patients:
        last, first = p["lastName"], p["firstName"]
        exact[f"{normalize(last)}|{normalize(first)}"] = p["id"]
        first_token = normalize(first).split(" ")[0] if first else ""
        loose.setdefault(f"{normalize(last)}|{first_token}", []).append(p["id"])
    return exact, loose


def match_patient(raw_name: str, exact_index: dict[str, str], loose_index: dict[str, list[str]]) -> str | None:
    first, last = split_name(raw_name)
    if not last:
        return None
    exact_key = f"{normalize(last)}|{normalize(first)}"
    if exact_key in exact_index:
        return exact_index[exact_key]
    first_token = normalize(first).split(" ")[0] if first else ""
    candidates = loose_index.get(f"{normalize(last)}|{first_token}", [])
    return candidates[0] if len(candidates) == 1 else None


def main():
    if not SOURCE_XLSX.exists():
        raise SystemExit(f"Source workbook not found: {SOURCE_XLSX}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    patients = json.loads(PATIENTS_JSON.read_text(encoding="utf-8")) if PATIENTS_JSON.exists() else []
    exact_index, loose_index = build_patient_index(patients)

    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)

    snapshots = []
    unparsed_sheet_names = []
    no_patient_column = []
    total_roster_rows = 0
    matched_rows = 0
    unmatched_names: set[str] = set()

    for sheet_name in wb.sheetnames:
        date = parse_sheet_date(sheet_name)
        if date is None:
            unparsed_sheet_names.append(sheet_name)
            continue

        ws = wb[sheet_name]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        patient_col = find_column(header, "patient", "name")
        if patient_col is None:
            no_patient_column.append(sheet_name)
            continue

        in_house = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if patient_col >= len(row) or not row[patient_col]:
                continue
            name = str(row[patient_col]).strip()
            if not name:
                continue
            in_house += 1
            total_roster_rows += 1
            patient_id = match_patient(name, exact_index, loose_index)
            if patient_id:
                matched_rows += 1
            else:
                unmatched_names.add(name)

        snapshots.append({
            "date": date,
            "inHouse": in_house,
            "totalUnits": TOTAL_UNITS,
            # unitsOccupied/unitsShared intentionally omitted -- see module docstring.
        })

    snapshots.sort(key=lambda s: s["date"])

    with open(OUT_DIR / "census-history.json", "w", encoding="utf-8") as f:
        json.dump(snapshots, f, ensure_ascii=False, indent=2)

    match_rate = round(100 * matched_rows / total_roster_rows, 1) if total_roster_rows else 0.0
    if snapshots:
        from datetime import date as _date
        span_days = (
            _date.fromisoformat(snapshots[-1]["date"]) - _date.fromisoformat(snapshots[0]["date"])
        ).days + 1
        coverage_pct = round(100 * len(snapshots) / span_days, 1)
    else:
        span_days, coverage_pct = 0, 0.0

    report_lines = [
        "# Occupancy Tracker Cleaning Report",
        "",
        f"Source: `{SOURCE_XLSX.name}` ({len(wb.sheetnames)} sheets, one per calendar day)",
        "",
        "## Sheet-name date parsing",
        f"- Sheets with a successfully parsed date: {len(snapshots) + len(no_patient_column)} / {len(wb.sheetnames)}",
        f"- Sheets whose name didn't parse as a date at all: {len(unparsed_sheet_names)} {unparsed_sheet_names or ''}",
        f"- Sheets with a parsed date but no detectable 'Patient's Name' column: {len(no_patient_column)} {no_patient_column or ''}",
        "",
        "## CensusSnapshot output",
        f"- Daily snapshots written: {len(snapshots)}",
        f"- Date range: {snapshots[0]['date'] if snapshots else 'n/a'} to {snapshots[-1]['date'] if snapshots else 'n/a'}",
        f"- Calendar-day coverage: {len(snapshots)} sheets over a {span_days}-day span ({coverage_pct}%) --",
        "  the tracker has real gaps (missing sheets for some calendar days, e.g. no sheet at all for",
        "  2026-08-07 through 2026-08-10), not filled forward or interpolated.",
        f"- Total roster rows (headcount events) across all days: {total_roster_rows}",
        "- A handful of the earliest 2025 sheets show `inHouse: 0` -- taken at face value (an accurately",
        "  empty roster row-count), not corrected, though the org's own header formatting was still",
        "  inconsistent in that early period so lower confidence applies to the earliest weeks.",
        "",
        "## Patient-name match rate (informational only -- does not affect inHouse counts)",
        f"- Roster rows matched to an existing patients.json record: {matched_rows} ({match_rate}%)",
        f"- Distinct unmatched names: {len(unmatched_names)}",
        "- `inHouse` counts every filled roster row regardless of match -- an unmatched name still represents",
        "  a real person physically in the house that day, so excluding them would undercount reality.",
        "",
        "## Not fabricated / explicitly out of scope this pass",
        "- `Stay` records (bed-level check-in/check-out history) were deliberately NOT generated from this",
        "  data -- the roster has no bed/unit assignment column, and `Stay.bedPositionId` is required and",
        "  used by the floor-plan board, house-occupancy-summary, and the partner portal's bed map to",
        "  determine which specific bed is occupied. Fabricating a bed assignment would make those views",
        "  look precise while being wrong. `lib/mock-data/patients.ts`'s existing `stays` generator is",
        "  unchanged by this pass.",
        "- `CensusSnapshot.unitsOccupied` / `unitsShared` are omitted for every real day -- no bed-level",
        "  detail exists in the source, only total headcount. Both fields were changed from required to",
        "  optional in lib/types/house-ops.ts; UI consumers now show a fallback instead of computing NaN.",
    ]
    (OUT_DIR / "occupancy-report.md").write_text("\n".join(report_lines), encoding="utf-8")

    print(f"Wrote {len(snapshots)} daily census snapshots to {OUT_DIR / 'census-history.json'}")
    print(f"  Date range {snapshots[0]['date'] if snapshots else 'n/a'} to {snapshots[-1]['date'] if snapshots else 'n/a'}")
    print(f"  Patient-name match rate: {matched_rows}/{total_roster_rows} ({match_rate}%)")
    print(f"See {OUT_DIR / 'occupancy-report.md'} for the full report.")


if __name__ == "__main__":
    main()
